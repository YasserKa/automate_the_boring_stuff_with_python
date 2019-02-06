import os
import openpyxl

dirPath = 'text_files'
wb = openpyxl.load_workbook('text_files.xlsx')
sheet = wb.get_active_sheet()

maxFiles = sheet.max_column
maxLines = sheet.max_row
#  print(sheet.column_dimensions['B'])
for column in range(1, maxFiles+1):
    fileStream = open(os.path.join(dirPath, str(column)), 'w')
    for row in range(1, maxLines):
        line = sheet.cell(row=row, column=column).value+'\n'
        fileStream.write(line)
    fileStream.close()
