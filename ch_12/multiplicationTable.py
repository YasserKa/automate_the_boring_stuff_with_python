#! /usr/bin/python3.6
import sys
import openpyxl
from openpyxl.styles import Font

N = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
# create the labels
for i in range(2, (N+2)):
    # generate row labels
    sheet.cell(row=i, column=1, value=i-1)
    # generate column labels
    sheet.cell(row=1, column=i, value=i-1)

boldFont = Font(bold=True)
sheet.row_dimensions[1].font = boldFont
sheet.column_dimensions['A'].font = boldFont

for column in range(2, (N+2)):
    for row in range(2, (N+2)):
        sheet.cell(
                    row=row, column=column,
                    value=sheet.cell(row=row, column=1).value
                    * sheet.cell(row=1, column=column).value
                )
wb.save('multi_table_by_'+str(N)+'.xlsx')
