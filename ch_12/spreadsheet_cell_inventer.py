import sys
import openpyxl

fileName = sys.argv[1]
wb = openpyxl.load_workbook(fileName)
sheet = wb.get_active_sheet()

highestRow = sheet.max_row
highestColumn = sheet.max_column

columnCount = 2
print(highestColumn+1)
for row in range(1, highestRow):
    for column in range(columnCount, highestColumn+1):
        cell1 = sheet.cell(row=row, column=column)
        cell2 = sheet.cell(row=column, column=row)
        tmp = cell1.value

        cell1.value = cell2.value
        cell2.value = tmp
    columnCount += 1

wb.save(fileName)
